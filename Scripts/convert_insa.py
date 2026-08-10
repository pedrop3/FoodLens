#!/usr/bin/env python3
"""Converte a Tabela da Composição de Alimentos do INSA (Excel, formato
PortFIR) para o JSON que o FoodIndex da app carrega em memória.

Uso:
    python3 convert_insa.py caminho/para/insa_tca.xlsx caminho/para/insa_foods.json

Espera a folha "INSA - BDCA_v X.Y - AAAA" (o nome exato muda a cada
atualização anual — o script procura a primeira folha cujo nome começa
por "INSA - BDCA", em vez de estar hardcoded à versão de hoje). Layout
fixo da tabela, linha 2 = cabeçalho, dados a partir da linha 3:
  coluna A = Cod, B = Nome do alimento, F = Energia [kcal] (por 100 g,
  exceto bebidas alcoólicas que são por 100 ml — ver nota abaixo).

Corre isto sempre que descarregares uma versão nova do Excel em
portfir.insa.pt; o output substitui Resources/insa_foods.json no projeto.
"""
import json
import sys
from pathlib import Path

import openpyxl

COD_COL = 0
NAME_COL = 1
KCAL_COL = 5


def find_data_sheet(workbook):
    for name in workbook.sheetnames:
        if name.startswith("INSA - BDCA"):
            return workbook[name]
    raise ValueError(
        f"Nenhuma folha 'INSA - BDCA...' encontrada. Folhas disponíveis: {workbook.sheetnames}"
    )


def convert(xlsx_path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = find_data_sheet(workbook)

    foods = []
    skipped = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        cod, name, kcal = row[COD_COL], row[NAME_COL], row[KCAL_COL]
        if cod is None or name is None:
            continue
        if not isinstance(kcal, (int, float)):
            # Tabela usa "" ou texto (ex.: valor não determinado) nalgumas
            # linhas para outros nutrientes; para energia não deve
            # acontecer, mas não vale a pena rebentar a conversão por isso
            # — melhor avisar e ignorar essa linha do que meter 0 kcal
            # silenciosamente (0 kcal errado é pior do que estar ausente).
            skipped.append((cod, name, kcal))
            continue
        foods.append({
            "id": str(cod).strip(),
            "name": str(name).strip(),
            "kcalPer100g": float(kcal),
        })

    if skipped:
        print(f"Aviso: {len(skipped)} linha(s) sem energia numérica, ignoradas:", file=sys.stderr)
        for cod, name, kcal in skipped[:20]:
            print(f"  - {cod} {name!r} energia={kcal!r}", file=sys.stderr)

    ids = [f["id"] for f in foods]
    if len(ids) != len(set(ids)):
        raise ValueError("IDs duplicados encontrados depois da conversão — verifica a coluna Cod.")

    return foods


def main() -> None:
    if len(sys.argv) != 3:
        print(f"Uso: {sys.argv[0]} <insa_tca.xlsx> <insa_foods.json>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    json_path = Path(sys.argv[2])

    foods = convert(xlsx_path)
    json_path.write_text(
        json.dumps(foods, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"{len(foods)} alimentos escritos em {json_path}")


if __name__ == "__main__":
    main()
